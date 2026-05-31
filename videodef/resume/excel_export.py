"""
Экспорт результатов поиска преподавателей в Excel с поэтапными вычислениями
Полная верификация метода Беллмана-Заде для реальных данных
"""

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from io import BytesIO
from datetime import datetime


class TeacherSearchExcelExporter:
    """
    Полноценный экспорт результатов поиска преподавателей в Excel
    с детальной поэтапной визуализацией всех вычислений
    """
    
    def __init__(self, model, results, alternatives_data, teacher_ids, 
                 user_location, criteria_weights, consistency_report, 
                 user_filters, use_expert_weights=True, 
                 criteria_comparisons=None, alternative_comparisons=None):
        
        self.model = model
        self.results = results
        self.alternatives_data = alternatives_data
        self.teacher_ids = teacher_ids
        self.user_location = user_location
        self.criteria_weights = criteria_weights
        self.consistency_report = consistency_report
        self.user_filters = user_filters
        self.use_expert_weights = use_expert_weights
        self.criteria_comparisons = criteria_comparisons or []
        self.alternative_comparisons = alternative_comparisons or {}
        
        # Полное описание критериев
        self.criteria_full = {
            'price': {'name': 'Цена занятия', 'unit': 'руб', 'direction': 'min', 'description': 'Стоимость одного академического часа'},
            'distance': {'name': 'Расстояние', 'unit': 'км', 'direction': 'min', 'description': 'Расстояние от места проживания до преподавателя'},
            'experience': {'name': 'Опыт работы', 'unit': 'лет', 'direction': 'max', 'description': 'Общий педагогический стаж'},
            'rating': {'name': 'Рейтинг', 'unit': 'звезд', 'direction': 'max', 'description': 'Средняя оценка от учеников'},
            'education': {'name': 'Уровень образования', 'unit': 'уровень', 'direction': 'max', 'description': 'Уровень профессионального образования'},
        }
        
        self.criteria_names = {k: v['name'] for k, v in self.criteria_full.items()}
        
        self._init_styles()
        
        self.alternatives = [f"T_{tid}" for tid in teacher_ids]
        self.alternative_names = {}
        for result in results:
            alt_id = f"T_{result['resume'].id}"
            self.alternative_names[alt_id] = self._get_teacher_name(result['resume'])
    
    def _get_teacher_name(self, resume):
        """Безопасное получение имени преподавателя"""
        try:
            if hasattr(resume.user, 'profile') and resume.user.profile:
                profile = resume.user.profile
                if profile.last_name and profile.first_name:
                    if profile.patronymic:
                        return f"{profile.last_name} {profile.first_name} {profile.patronymic}"
                    return f"{profile.last_name} {profile.first_name}"
        except Exception:
            pass
        return f"Преподаватель {resume.id}"
    
    def _init_styles(self):
        """Инициализация стилей Excel"""
        self.title_font = Font(bold=True, size=14, color="FFFFFF")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.subheader_font = Font(bold=True, size=12)
        self.bold_font = Font(bold=True)
        self.formula_font = Font(bold=True, italic=True, color="0000FF")
        
        self.title_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        self.header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        self.success_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        self.best_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        self.formula_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        self.thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    def _get_col_letter(self, col):
        """Преобразование номера колонки в букву"""
        result = ""
        while col > 0:
            col -= 1
            result = chr(ord('A') + col % 26) + result
            col //= 26
        return result
    
    # ==================== ЛИСТ 0: ОБЩАЯ ИНФОРМАЦИЯ ====================
    
    def create_info_sheet(self, wb):
        """Лист 0: Общая информация о поиске"""
        ws = wb.create_sheet("0_Общая_информация")
        
        ws.merge_cells('A1:F1')
        cell = ws.cell(row=1, column=1, value="ОТЧЕТ ПО РЕЗУЛЬТАТАМ ПОИСКА ПРЕПОДАВАТЕЛЕЙ")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        meta_data = [
            ("Дата формирования отчета:", datetime.now().strftime("%d.%m.%Y %H:%M:%S")),
            ("Метод анализа:", "Нечеткий многокритериальный анализ Беллмана-Заде"),
            ("Версия метода:", "Полная реализация с парными сравнениями (матрицы Саати)"),
            ("Источник весов:", "ЭКСПЕРТНЫЕ" if self.use_expert_weights else "ПОЛЬЗОВАТЕЛЬСКИЕ"),
            ("Количество альтернатив:", str(len(self.results))),
            ("Количество критериев:", str(len(self.model.criteria))),
        ]
        
        row = 3
        for label, value in meta_data:
            ws.cell(row=row, column=1, value=label).font = self.bold_font
            ws.cell(row=row, column=2, value=value)
            if "весов" in label.lower():
                if self.use_expert_weights:
                    ws.cell(row=row, column=2).fill = self.success_fill
                else:
                    ws.cell(row=row, column=2).fill = self.highlight_fill
            row += 1
        
        # Параметры фильтрации
        row += 1
        header_cell = ws.cell(row=row, column=1, value="ПАРАМЕТРЫ ПОИСКА")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = self.subheader_fill
        
        row += 1
        filters_data = [
            ("Диапазон цен:", f"{self.user_filters.get('price_min', 0):.0f} руб - {self.user_filters.get('price_max', 10000):.0f} руб"),
            ("Максимальное расстояние:", f"{self.user_filters.get('max_distance', 20):.1f} км"),
            ("Минимальный опыт:", f"{self.user_filters.get('min_experience', 0):.1f} лет"),
            ("Минимальный рейтинг:", f"{self.user_filters.get('min_rating', 0):.1f} звезд"),
            ("Режим весов:", "ЭКСПЕРТНЫЙ" if self.use_expert_weights else "ПОЛЬЗОВАТЕЛЬСКИЙ"),
        ]
        
        for label, value in filters_data:
            ws.cell(row=row, column=1, value=label).font = self.bold_font
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Местоположение
        row += 1
        header_cell = ws.cell(row=row, column=1, value="МЕСТОПОЛОЖЕНИЕ ПОЛЬЗОВАТЕЛЯ")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = self.subheader_fill
        
        row += 1
        ws.cell(row=row, column=1, value="Широта:").font = self.bold_font
        ws.cell(row=row, column=2, value=f"{self.user_location[0]:.6f}°")
        row += 1
        ws.cell(row=row, column=1, value="Долгота:").font = self.bold_font
        ws.cell(row=row, column=2, value=f"{self.user_location[1]:.6f}°")
        row += 1
        ws.cell(row=row, column=1, value="Метод определения:").font = self.bold_font
        ws.cell(row=row, column=2, value="Из профиля пользователя / ребенка / по умолчанию")
        
        # Краткий обзор
        row += 2
        header_cell = ws.cell(row=row, column=1, value="КРАТКИЙ ОБЗОР РЕЗУЛЬТАТОВ")
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = self.subheader_fill
        
        row += 1
        if self.results:
            best = self.results[0]
            best_name = self._get_teacher_name(best['resume'])
            
            ws.cell(row=row, column=1, value="Лучший преподаватель:").font = self.bold_font
            ws.cell(row=row, column=2, value=best_name)
            ws.cell(row=row, column=3, value=f"μD = {best['mu']:.2f}%")
            row += 1
            
            ws.cell(row=row, column=1, value="Ранжирование ТОП-3:").font = self.bold_font
            ranking_parts = []
            for r, res in enumerate(self.results[:3]):
                name = self._get_teacher_name(res['resume'])
                ranking_parts.append(f"{r+1}. {name[:25]} ({res['mu']:.1f}%)")
            ranking_str = "  ->  ".join(ranking_parts)
            ws.cell(row=row, column=2, value=ranking_str)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
        
        return ws
    
    # ==================== ЛИСТ 1: ИСХОДНЫЕ ДАННЫЕ ====================
    
    def create_raw_data_sheet(self, wb):
        """Лист 1: Исходные данные преподавателей"""
        ws = wb.create_sheet("1_Исходные_данные")
        
        ws.merge_cells('A1:M1')
        cell = ws.cell(row=1, column=1, value="ИСХОДНЫЕ ДАННЫЕ ПРЕПОДАВАТЕЛЕЙ")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Примечание:").font = self.bold_font
        ws.cell(row=2, column=2, value="Данные нормализованы по шкале от 1 до 10 для построения матриц парных сравнений")
        ws.cell(row=2, column=2).fill = self.formula_fill
        
        # Формулы нормализации
        ws.cell(row=3, column=1, value="Формулы нормализации:").font = self.bold_font
        ws.cell(row=3, column=2, value="Цена (min): норма = 10 - (цена / 10000) * 9")
        ws.cell(row=4, column=2, value="Расстояние (min): норма = 10 - (расстояние / 50) * 9")
        ws.cell(row=5, column=2, value="Опыт (max): норма = 1 + (опыт / 30) * 9")
        ws.cell(row=6, column=2, value="Рейтинг (max): норма = 1 + (рейтинг / 5) * 9")
        ws.cell(row=7, column=2, value="Образование (max): норма = 1 + (уровень / 10) * 9")
        
        headers = ["ID", "Преподаватель", "Цена (руб)", "Расстояние (км)", 
                   "Опыт (лет)", "Рейтинг", "Образование",
                   "Норм. цена", "Норм. расстояние", "Норм. опыт", 
                   "Норм. рейтинг", "Норм. образование", "Источник координат"]
        
        row = 9
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row = 10
        for result in self.results:
            resume = result['resume']
            teacher_name = self._get_teacher_name(resume)
            alt_id = f"T_{resume.id}"
            alt_data = result.get('raw_data', {})
            
            # Получаем нормализованные значения
            normalized = {}
            for criterion in self.model.criteria:
                fuzzy_set = self.model.criterion_fuzzy_sets.get(criterion)
                if fuzzy_set and alt_id in fuzzy_set.memberships:
                    normalized[criterion] = fuzzy_set.memberships[alt_id] * 10
                else:
                    normalized[criterion] = 5
            
            location_source = "Из профиля" if resume.location_lat else "Не указан"
            
            ws.cell(row=row, column=1, value=resume.id).alignment = self.center_align
            ws.cell(row=row, column=2, value=teacher_name).alignment = self.left_align
            ws.cell(row=row, column=3, value=round(alt_data.get('price', 0), 2)).alignment = self.center_align
            ws.cell(row=row, column=4, value=round(alt_data.get('distance', 0), 3)).alignment = self.center_align
            ws.cell(row=row, column=5, value=alt_data.get('experience', 0)).alignment = self.center_align
            ws.cell(row=row, column=6, value=alt_data.get('rating', 0)).alignment = self.center_align
            ws.cell(row=row, column=7, value=alt_data.get('education', 0)).alignment = self.center_align
            ws.cell(row=row, column=8, value=round(normalized.get('price', 5), 2)).alignment = self.center_align
            ws.cell(row=row, column=9, value=round(normalized.get('distance', 5), 2)).alignment = self.center_align
            ws.cell(row=row, column=10, value=round(normalized.get('experience', 5), 2)).alignment = self.center_align
            ws.cell(row=row, column=11, value=round(normalized.get('rating', 5), 2)).alignment = self.center_align
            ws.cell(row=row, column=12, value=round(normalized.get('education', 5), 2)).alignment = self.center_align
            ws.cell(row=row, column=13, value=location_source).alignment = self.center_align
            
            for col in range(1, 14):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Условное форматирование
        for col in range(8, 13):
            col_letter = self._get_col_letter(col)
            if row > 10:
                ws.conditional_formatting.add(
                    f'{col_letter}10:{col_letter}{row-1}',
                    ColorScaleRule(start_type='min', start_color='FF0000',
                                  mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                  end_type='max', end_color='00FF00')
                )
        
        widths = [8, 28, 12, 14, 10, 10, 12, 11, 14, 11, 12, 14, 20]
        for i, width in enumerate(widths, 1):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = width
        
        return ws
    
    # ==================== ЛИСТ 2: ЭКСПЕРТНЫЕ СРАВНЕНИЯ ====================
    
    def create_expert_comparisons_sheet(self, wb):
        """Лист 2: Экспертные парные сравнения"""
        ws = wb.create_sheet("2_Экспертные_сравнения")
        
        ws.merge_cells('A1:E1')
        cell = ws.cell(row=1, column=1, value="ЭКСПЕРТНЫЕ ПАРНЫЕ СРАВНЕНИЯ (ШКАЛА СААТИ)")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        # Описание шкалы Саати
        ws.cell(row=2, column=1, value="Шкала относительной важности Саати:").font = self.bold_font
        saaty_scale = [
            "1 - Одинаковая важность",
            "2 - Почти слабое преимущество",
            "3 - Слабое преимущество",
            "4 - Почти существенное преимущество",
            "5 - Существенное преимущество",
            "6 - Почти сильное преимущество",
            "7 - Явное преимущество",
            "8 - Очень сильное преимущество",
            "9 - Абсолютное преимущество"
        ]
        
        row = 3
        for desc in saaty_scale:
            ws.cell(row=row, column=1, value=desc)
            row += 1
        
        # Сравнения критериев
        if self.criteria_comparisons:
            row += 2
            header_cell = ws.cell(row=row, column=1, value="СРАВНЕНИЯ КРИТЕРИЕВ")
            header_cell.font = Font(bold=True, color="FFFFFF")
            header_cell.fill = self.subheader_fill
            
            row += 1
            headers = ["Критерий 1", "Критерий 2", "Лингвистическая оценка", "Числовое значение", "Интерпретация"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align
                cell.border = self.border
            
            row += 1
            for comp in self.criteria_comparisons:
                c1 = comp.get('criterion1', '')
                c2 = comp.get('criterion2', '')
                ling = comp.get('linguistic_value', '')
                value = comp.get('value', 1)
                
                if isinstance(value, str) and '/' in value:
                    num, den = value.split('/')
                    value = float(num) / float(den)
                else:
                    try:
                        value = float(value)
                    except:
                        value = 1
                
                if value > 1:
                    interpretation = f"{self.criteria_names.get(c1, c1)} важнее {self.criteria_names.get(c2, c2)}"
                elif value < 1:
                    interpretation = f"{self.criteria_names.get(c2, c2)} важнее {self.criteria_names.get(c1, c1)}"
                else:
                    interpretation = "Критерии равнозначны"
                
                ws.cell(row=row, column=1, value=self.criteria_names.get(c1, c1)).alignment = self.center_align
                ws.cell(row=row, column=2, value=self.criteria_names.get(c2, c2)).alignment = self.center_align
                ws.cell(row=row, column=3, value=ling).alignment = self.center_align
                ws.cell(row=row, column=4, value=round(value, 3)).alignment = self.center_align
                ws.cell(row=row, column=5, value=interpretation).alignment = self.left_align
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.border
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 45
        
        return ws
    
    # ==================== ЛИСТ 3: МАТРИЦЫ ПАРНЫХ СРАВНЕНИЙ ====================
    
    def create_matrices_sheet(self, wb):
        """Лист 3: Матрицы парных сравнений по критериям"""
        ws = wb.create_sheet("3_Матрицы_сравнений")
        
        max_col = len(self.alternatives) + 2
        end_col = self._get_col_letter(max_col)
        ws.merge_cells(f'A1:{end_col}1')
        cell = ws.cell(row=1, column=1, value="МАТРИЦЫ ПАРНЫХ СРАВНЕНИЙ АЛЬТЕРНАТИВ")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Метод построения:")
        ws.cell(row=2, column=2, value="Автоматическое на основе нормализованных данных (отношение значений)")
        ws.cell(row=3, column=1, value="Формула:")
        ws.cell(row=3, column=2, value="a_ij = значение_альтернативы_i / значение_альтернативы_j, приведенное к шкале 1/9-9")
        
        row = 6
        for criterion in self.model.criteria:
            matrix = self.model.criterion_matrices.get(criterion)
            if not matrix:
                continue
            
            # Заголовок матрицы
            ws.merge_cells(f'A{row}:{end_col}{row}')
            cell = ws.cell(row=row, column=1, value=f"МАТРИЦА: {self.criteria_names.get(criterion, criterion)}")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = self.subheader_fill
            cell.alignment = self.center_align
            row += 1
            
            # Заголовки альтернатив
            for col, alt in enumerate(self.alternatives, 2):
                cell = ws.cell(row=row, column=col, value=self.alternative_names.get(alt, alt)[:20])
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align
                cell.border = self.border
            
            # Строки матрицы
            for i, alt1 in enumerate(self.alternatives):
                ws.cell(row=row+1+i, column=1, value=self.alternative_names.get(alt1, alt1)[:20])
                ws.cell(row=row+1+i, column=1).font = self.header_font
                ws.cell(row=row+1+i, column=1).fill = self.header_fill
                ws.cell(row=row+1+i, column=1).alignment = self.center_align
                ws.cell(row=row+1+i, column=1).border = self.border
                
                for j, alt2 in enumerate(self.alternatives):
                    val = matrix.get_comparison(i, j)
                    cell = ws.cell(row=row+1+i, column=2+j, value=round(val, 3))
                    cell.alignment = self.center_align
                    cell.border = self.border
                    
                    if i == j:
                        cell.fill = self.formula_fill
                    elif i < j and val >= 1:
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            row += len(self.alternatives) + 3
        
        ws.column_dimensions['A'].width = 28
        for i in range(2, len(self.alternatives) + 3):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = 22
        
        return ws
    
    # ==================== ЛИСТ 4: НЕЧЕТКИЕ МНОЖЕСТВА ====================
    
    def create_fuzzy_sets_sheet(self, wb):
        """Лист 4: Нечеткие множества (степени принадлежности)"""
        ws = wb.create_sheet("4_Нечеткие_множества")
        
        ws.merge_cells('A1:G1')
        cell = ws.cell(row=1, column=1, value="НЕЧЕТКИЕ МНОЖЕСТВА (СТЕПЕНИ ПРИНАДЛЕЖНОСТИ)")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Формула расчета:")
        ws.cell(row=2, column=2, value="Степени принадлежности = нормированный собственный вектор матрицы парных сравнений")
        ws.cell(row=2, column=2).fill = self.formula_fill
        
        ws.cell(row=3, column=1, value="Математическое обоснование:")
        ws.cell(row=3, column=2, value="Aw = λ_max·w, где w - собственный вектор, λ_max - максимальное собственное значение")
        
        ws.cell(row=4, column=1, value="Интерпретация:")
        ws.cell(row=4, column=2, value="μ ∈ [0, 1]. Чем выше значение, тем лучше альтернатива по данному критерию")
        
        # Таблица принадлежностей
        row = 6
        headers = ["Критерий", "Альтернатива", "Преподаватель", "Степень принадлежности μ", 
                   "Норм. значение (1-10)", "Ранг", "Качество"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row = 7
        
        for criterion in self.model.criteria:
            fuzzy_set = self.model.criterion_fuzzy_sets.get(criterion)
            if not fuzzy_set:
                continue
            
            ranked = fuzzy_set.get_ranked_elements()
            criterion_name = self.criteria_names.get(criterion, criterion)
            
            for rank, (alt, mu) in enumerate(ranked, 1):
                alt_name = self.alternative_names.get(alt, alt)
                norm_val = mu * 10
                
                if mu >= 0.7:
                    quality = "Высокий"
                elif mu >= 0.4:
                    quality = "Средний"
                else:
                    quality = "Низкий"
                
                ws.cell(row=row, column=1, value=criterion_name).alignment = self.center_align
                ws.cell(row=row, column=2, value=alt).alignment = self.center_align
                ws.cell(row=row, column=3, value=alt_name).alignment = self.left_align
                ws.cell(row=row, column=4, value=round(mu, 4)).alignment = self.center_align
                ws.cell(row=row, column=5, value=round(norm_val, 2)).alignment = self.center_align
                ws.cell(row=row, column=6, value=rank).alignment = self.center_align
                ws.cell(row=row, column=7, value=quality).alignment = self.center_align
                
                if rank == 1:
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = self.best_fill
                
                if quality == "Низкий":
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
            row += 1
        
        widths = [20, 12, 28, 18, 18, 8, 12]
        for i, width in enumerate(widths, 1):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = width
        
        return ws
    
    # ==================== ЛИСТ 5: ВЕСА КРИТЕРИЕВ ====================
    
    def create_weights_sheet(self, wb):
        """Лист 5: Веса критериев (α-коэффициенты)"""
        ws = wb.create_sheet("5_Веса_критериев")
        
        ws.merge_cells('A1:H1')
        cell = ws.cell(row=1, column=1, value="ВЕСА КРИТЕРИЕВ (α-КОЭФФИЦИЕНТЫ)")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Источник весов:")
        if self.use_expert_weights:
            ws.cell(row=2, column=2, value="ЭКСПЕРТНЫЕ ОЦЕНКИ (из настроек администратора)")
            ws.cell(row=2, column=2).fill = self.success_fill
        else:
            ws.cell(row=2, column=2, value="ПОЛЬЗОВАТЕЛЬСКИЕ НАСТРОЙКИ")
            ws.cell(row=2, column=2).fill = self.highlight_fill
        
        ws.cell(row=3, column=1, value="Метод расчета:")
        ws.cell(row=3, column=2, value="Метод собственного вектора (нормализация) - Aw = λ_max·w")
        
        ws.cell(row=4, column=1, value="Условие нормировки:")
        ws.cell(row=4, column=2, value="Σ αⱼ = 1")
        
        # Таблица весов
        row = 6
        headers = ["Критерий", "Название", "Цель", "Единица", "Вес (α)", "Вес (%)", "Важность", "Влияние"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row = 7
        
        sorted_criteria = sorted(self.model.criteria, 
                                 key=lambda c: self.criteria_weights.get(c, 0), 
                                 reverse=True)
        
        for criterion in sorted_criteria:
            weight = self.criteria_weights.get(criterion, 0)
            info = self.criteria_full.get(criterion, {})
            goal = "минимизация" if info.get('direction') == 'min' else "максимизация"
            unit = info.get('unit', '')
            
            if weight > 0.25:
                importance = "Очень высокая"
            elif weight > 0.15:
                importance = "Высокая"
            elif weight > 0.1:
                importance = "Средняя"
            else:
                importance = "Низкая"
            
            influence = f"{(weight * 100):.1f}% решения"
            
            ws.cell(row=row, column=1, value=criterion).alignment = self.center_align
            ws.cell(row=row, column=2, value=info.get('name', criterion)).alignment = self.left_align
            ws.cell(row=row, column=3, value=goal).alignment = self.center_align
            ws.cell(row=row, column=4, value=unit).alignment = self.center_align
            ws.cell(row=row, column=5, value=round(weight, 4)).alignment = self.center_align
            ws.cell(row=row, column=6, value=f"{weight*100:.1f}%").alignment = self.center_align
            ws.cell(row=row, column=7, value=importance).alignment = self.center_align
            ws.cell(row=row, column=8, value=influence).alignment = self.center_align
            
            if weight > 0.25:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Формула свертки
        row += 2
        ws.cell(row=row, column=1, value="ФОРМУЛА СВЕРТКИ (агрегирования):").font = self.bold_font
        row += 1
        cell = ws.cell(row=row, column=1, value="μD(P) = min( μ₁(P)^α₁ , μ₂(P)^α₂ , ... , μₙ(P)^αₙ )")
        cell.font = self.formula_font
        cell.fill = self.formula_fill
        
        widths = [10, 25, 14, 10, 12, 12, 16, 15]
        for i, width in enumerate(widths, 1):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = width
        
        return ws
    
    # ==================== ЛИСТ 6: ПОШАГОВЫЙ РАСЧЕТ μD ====================
    
    def create_step_by_step_sheet(self, wb):
        """Лист 6: Пошаговый расчет μD для каждой альтернативы"""
        ws = wb.create_sheet("6_Пошаговый_расчет_μD")
        
        ws.merge_cells('A1:H1')
        cell = ws.cell(row=1, column=1, value="ПОШАГОВЫЙ РАСЧЕТ СТЕПЕНИ ПРИНАДЛЕЖНОСТИ μD")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Формула Беллмана-Заде (2.35 из методички):")
        ws.cell(row=2, column=2, value="μD(P) = min( μ₁(P)^α₁, μ₂(P)^α₂, ..., μₙ(P)^αₙ )")
        ws.cell(row=2, column=2).font = self.formula_font
        
        ws.cell(row=3, column=1, value="Алгоритм вычисления:")
        steps = [
            "1. Для каждого критерия вычисляем степень принадлежности μⱼ(P)",
            "2. Возводим μⱼ(P) в степень αⱼ (вес критерия)",
            "3. Находим минимальное значение среди всех μⱼ(P)^αⱼ",
            "4. Это минимальное значение и есть μD(P)"
        ]
        
        row = 4
        for step in steps:
            ws.cell(row=row, column=1, value=step)
            row += 1
        
        row = 7
        
        for result in self.results:
            resume = result['resume']
            teacher_name = self._get_teacher_name(resume)
            
            # Заголовок альтернативы
            ws.merge_cells(f'A{row}:H{row}')
            cell = ws.cell(row=row, column=1, value=f"АЛЬТЕРНАТИВА: {teacher_name}")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = self.subheader_fill
            cell.alignment = self.center_align
            row += 1
            
            # Заголовки таблицы
            headers = ["Критерий", "Исходное значение", "μ", "α", "μ^α", "Промежуточный min", "Примечание"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_align
                cell.border = self.border
            row += 1
            
            weighted_values = []
            min_so_far = 1.0
            min_criterion = None
            
            for criterion in self.model.criteria:
                mu = result['memberships'].get(criterion, 0) / 100
                alpha = self.criteria_weights.get(criterion, 0.2)
                mu_alpha = mu ** alpha if mu > 0 else 0
                weighted_values.append(mu_alpha)
                
                if mu_alpha < min_so_far:
                    min_so_far = mu_alpha
                    min_criterion = criterion
                
                raw_value = result.get(criterion, result['raw_data'].get(criterion, 0))
                if criterion == 'price':
                    raw_str = f"{raw_value:.0f} руб"
                elif criterion == 'distance':
                    raw_str = f"{raw_value:.2f} км"
                elif criterion == 'experience':
                    raw_str = f"{raw_value:.1f} лет"
                elif criterion == 'rating':
                    raw_str = f"{raw_value:.1f} звезд"
                else:
                    raw_str = str(raw_value)
                
                ws.cell(row=row, column=1, value=self.criteria_names.get(criterion, criterion)).alignment = self.left_align
                ws.cell(row=row, column=2, value=raw_str).alignment = self.center_align
                ws.cell(row=row, column=3, value=round(mu, 4)).alignment = self.center_align
                ws.cell(row=row, column=4, value=round(alpha, 4)).alignment = self.center_align
                ws.cell(row=row, column=5, value=round(mu_alpha, 4)).alignment = self.center_align
                
                if mu_alpha == min_so_far:
                    ws.cell(row=row, column=6, value=round(min_so_far, 4)).alignment = self.center_align
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).fill = self.highlight_fill
                
                note = ""
                if mu_alpha == min_so_far:
                    note = "Текущий минимум (определяет μD)"
                elif mu_alpha < 0.3:
                    note = "Низкое значение"
                elif mu_alpha > 0.8:
                    note = "Высокое значение"
                
                ws.cell(row=row, column=7, value=note).alignment = self.left_align
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
            
            # Итоговая строка
            mu_d = result['mu'] / 100
            min_val = min(weighted_values) if weighted_values else 0
            
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            cell = ws.cell(row=row, column=1, value="ИТОГОВОЕ μD = min(μ^α)")
            cell.font = self.bold_font
            cell.fill = self.success_fill
            
            ws.cell(row=row, column=3, value=f"= {min_val:.4f}").font = self.bold_font
            ws.cell(row=row, column=4, value=f"= {mu_d:.4f}").font = self.bold_font
            ws.cell(row=row, column=5, value=f"{result['mu']:.2f}%").font = self.bold_font
            ws.cell(row=row, column=6, value=f"Минимум по '{self.criteria_names.get(min_criterion, min_criterion)}'").alignment = self.left_align
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thick_border
            
            row += 3
        
        widths = [22, 18, 10, 10, 12, 15, 35]
        for i, width in enumerate(widths, 1):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = width
        
        return ws
    
    # ==================== ЛИСТ 7: ФИНАЛЬНОЕ РАНЖИРОВАНИЕ ====================
    
    def create_final_results_sheet(self, wb):
        """Лист 7: Финальное ранжирование преподавателей"""
        ws = wb.create_sheet("7_Финальное_ранжирование")
        
        ws.merge_cells('A1:L1')
        cell = ws.cell(row=1, column=1, value="ФИНАЛЬНОЕ РАНЖИРОВАНИЕ ПРЕПОДАВАТЕЛЕЙ")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Принцип принятия решения:")
        ws.cell(row=2, column=2, value="Выбирается альтернатива с МАКСИМАЛЬНОЙ степенью принадлежности μD")
        
        ws.cell(row=3, column=1, value="Интерпретация μD:")
        interpretations = [
            "90-100%: Идеальное соответствие",
            "70-89%: Высокое соответствие",
            "50-69%: Среднее соответствие",
            "30-49%: Низкое соответствие",
            "0-29%: Не соответствует требованиям"
        ]
        
        row = 4
        for interp in interpretations:
            ws.cell(row=row, column=1, value=interp)
            row += 1
        
        # Таблица результатов
        row = 7
        headers = ["Ранг", "Преподаватель", "μD (%)", "Цена (руб)", "Расстояние (км)", 
                   "Опыт (лет)", "Рейтинг", "Лучшие критерии", "Слабые критерии", 
                   "Рекомендация", "Действие"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row = 8
        
        for result in self.results:
            resume = result['resume']
            teacher_name = self._get_teacher_name(resume)
            memberships = result.get('memberships', {})
            
            strong_criteria = []
            weak_criteria = []
            
            for criterion, mu in memberships.items():
                if mu >= 70:
                    strong_criteria.append(self.criteria_names.get(criterion, criterion)[:2])
                elif mu <= 30:
                    weak_criteria.append(self.criteria_names.get(criterion, criterion)[:2])
            
            mu = result['mu']
            if result['rank'] == 1:
                recommendation = "РЕКОМЕНДУЕТСЯ"
                action = "Связаться с преподавателем"
            elif mu >= 70:
                recommendation = "Хороший вариант"
                action = "Рассмотреть для пробного занятия"
            elif mu >= 50:
                recommendation = "Возможный вариант"
                action = "Изучить детали профиля"
            else:
                recommendation = "Не рекомендуется"
                action = "Продолжить поиск"
            
            ws.cell(row=row, column=1, value=result['rank']).alignment = self.center_align
            ws.cell(row=row, column=2, value=teacher_name).alignment = self.left_align
            ws.cell(row=row, column=3, value=round(mu, 2)).alignment = self.center_align
            ws.cell(row=row, column=4, value=round(result.get('price', 0), 2)).alignment = self.center_align
            ws.cell(row=row, column=5, value=round(result.get('distance', 0), 2)).alignment = self.center_align
            ws.cell(row=row, column=6, value=result.get('experience', 0)).alignment = self.center_align
            ws.cell(row=row, column=7, value=result.get('rating', 0)).alignment = self.center_align
            ws.cell(row=row, column=8, value=", ".join(strong_criteria) if strong_criteria else "—").alignment = self.center_align
            ws.cell(row=row, column=9, value=", ".join(weak_criteria) if weak_criteria else "—").alignment = self.center_align
            ws.cell(row=row, column=10, value=recommendation).alignment = self.center_align
            ws.cell(row=row, column=11, value=action).alignment = self.center_align
            
            if result['rank'] == 1:
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = self.best_fill
            
            if mu >= 70:
                ws.cell(row=row, column=3).fill = self.success_fill
            elif mu <= 30:
                ws.cell(row=row, column=3).fill = self.warning_fill
            
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Гистограмма распределения μD
        row += 2
        ws.cell(row=row, column=1, value="ГИСТОГРАММА РАСПРЕДЕЛЕНИЯ μD:").font = self.bold_font
        row += 1
        
        max_mu = max([r['mu'] for r in self.results]) if self.results else 100
        for result in self.results:
            teacher_name = self._get_teacher_name(result['resume'])[:20]
            mu = result['mu']
            bar_length = int(mu / (max_mu / 30)) if max_mu > 0 else 0
            bar = "█" * min(bar_length, 40) + "░" * max(0, 40 - bar_length)
            
            ws.cell(row=row, column=1, value=teacher_name)
            ws.cell(row=row, column=2, value=bar)
            ws.cell(row=row, column=3, value=f"{mu:.1f}%")
            row += 1
        
        widths = [25, 35, 10, 12, 14, 10, 10, 15, 15, 22, 25]
        for i, width in enumerate(widths, 1):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = width
        
        return ws
    
    # ==================== ЛИСТ 8: СОГЛАСОВАННОСТЬ МАТРИЦ ====================
    
    def create_consistency_sheet(self, wb):
        """Лист 8: Оценка согласованности матриц"""
        ws = wb.create_sheet("8_Согласованность_матриц")
        
        ws.merge_cells('A1:G1')
        cell = ws.cell(row=1, column=1, value="ОЦЕНКА СОГЛАСОВАННОСТИ МАТРИЦ ПАРНЫХ СРАВНЕНИЙ")
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_align
        
        ws.cell(row=2, column=1, value="Критический уровень:")
        ws.cell(row=2, column=2, value="CR (Consistency Ratio) должен быть < 0.1 (10%)")
        
        ws.cell(row=3, column=1, value="Формулы расчета:")
        ws.cell(row=3, column=2, value="λ_max - максимальное собственное значение матрицы")
        ws.cell(row=4, column=2, value="CI = (λ_max - n) / (n - 1) - индекс согласованности")
        ws.cell(row=5, column=2, value="CR = CI / RI - отношение согласованности")
        ws.cell(row=6, column=2, value="RI - случайный индекс (зависит от размера матрицы n)")
        
        # Таблица случайных индексов
        row = 8
        ws.cell(row=row, column=1, value="Случайные индексы (RI):").font = self.bold_font
        row += 1
        
        ri_values = [("n=1", "0.00"), ("n=2", "0.00"), ("n=3", "0.58"), ("n=4", "0.90"),
                     ("n=5", "1.12"), ("n=6", "1.24"), ("n=7", "1.32"), ("n=8", "1.41"),
                     ("n=9", "1.45"), ("n=10", "1.49")]
        
        for i, (n, ri) in enumerate(ri_values):
            ws.cell(row=row, column=i+1, value=n).alignment = self.center_align
            ws.cell(row=row, column=i+1).font = self.header_font
            ws.cell(row=row, column=i+1).fill = self.header_fill
            ws.cell(row=row+1, column=i+1, value=ri).alignment = self.center_align
            ws.cell(row=row+1, column=i+1).border = self.border
        
        # Таблица согласованности
        row += 4
        headers = ["Матрица", "Размерность", "RI", "λ_max", "CI", "CR", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        row += 1
        
        # Матрица критериев
        if self.consistency_report and self.consistency_report.get('criteria_matrix'):
            cr_data = self.consistency_report['criteria_matrix']
            cr = cr_data.get('cr', 1)
            n = len(self.model.criteria)
            ri_map = {3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41}
            ri = ri_map.get(n, 1.24)
            
            status = "СОГЛАСОВАНА" if cr < 0.1 else "НЕСОГЛАСОВАНА"
            
            ws.cell(row=row, column=1, value="Критерии").alignment = self.center_align
            ws.cell(row=row, column=2, value=n).alignment = self.center_align
            ws.cell(row=row, column=3, value=ri).alignment = self.center_align
            ws.cell(row=row, column=4, value=round(cr_data.get('lambda_max', 0), 4)).alignment = self.center_align
            ws.cell(row=row, column=5, value=round(cr_data.get('ci', 0), 4)).alignment = self.center_align
            ws.cell(row=row, column=6, value=round(cr, 4)).alignment = self.center_align
            ws.cell(row=row, column=7, value=status).alignment = self.center_align
            
            if cr < 0.1:
                ws.cell(row=row, column=7).fill = self.success_fill
            else:
                ws.cell(row=row, column=7).fill = self.warning_fill
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.border
            row += 1
        
        # Матрицы по критериям
        if self.consistency_report and self.consistency_report.get('criterion_matrices'):
            for criterion, cr_data in self.consistency_report['criterion_matrices'].items():
                cr = cr_data.get('cr', 1)
                n = len(self.alternatives)
                ri_map = {3: 0.58, 4: 0.90, 5: 1.12}
                ri = ri_map.get(n, 0.90)
                
                status = "СОГЛАСОВАНА" if cr < 0.1 else "НЕСОГЛАСОВАНА"
                
                ws.cell(row=row, column=1, value=self.criteria_names.get(criterion, criterion)).alignment = self.center_align
                ws.cell(row=row, column=2, value=n).alignment = self.center_align
                ws.cell(row=row, column=3, value=ri).alignment = self.center_align
                ws.cell(row=row, column=4, value=round(cr_data.get('lambda_max', 0), 4)).alignment = self.center_align
                ws.cell(row=row, column=5, value=round(cr_data.get('ci', 0), 4)).alignment = self.center_align
                ws.cell(row=row, column=6, value=round(cr, 4)).alignment = self.center_align
                ws.cell(row=row, column=7, value=status).alignment = self.center_align
                
                if cr < 0.1:
                    ws.cell(row=row, column=7).fill = self.success_fill
                else:
                    ws.cell(row=row, column=7).fill = self.warning_fill
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.border
                row += 1
        
        # Рекомендации
        row += 2
        ws.cell(row=row, column=1, value="РЕКОМЕНДАЦИИ:").font = self.bold_font
        row += 1
        recommendations = [
            "1. Матрицы с CR > 0.1 считаются несогласованными",
            "2. Несогласованность может привести к искажению результатов анализа",
            "3. Для корректировки необходимо изменить противоречивые парные сравнения",
            "4. Рекомендуется использовать шкалу Саати для оценки важности",
            "5. При большом количестве альтернатив (>7) допустим CR < 0.15"
        ]
        
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec)
            row += 1
        
        for i in range(1, 8):
            col_letter = self._get_col_letter(i)
            ws.column_dimensions[col_letter].width = 18
        
        return ws
    
    # ==================== ЭКСПОРТ ====================
    
    def export(self, filename):
        """Экспорт всех данных в Excel файл"""
        wb = Workbook()
        wb.remove(wb.active)
        
        self.create_info_sheet(wb)                # Лист 0
        self.create_raw_data_sheet(wb)            # Лист 1
        self.create_expert_comparisons_sheet(wb)  # Лист 2
        self.create_matrices_sheet(wb)            # Лист 3
        self.create_fuzzy_sets_sheet(wb)          # Лист 4
        self.create_weights_sheet(wb)             # Лист 5
        self.create_step_by_step_sheet(wb)        # Лист 6
        self.create_final_results_sheet(wb)       # Лист 7
        self.create_consistency_sheet(wb)         # Лист 8
        
        wb.save(filename)
        return filename
    
    def export_to_bytesio(self):
        """Экспорт в BytesIO для передачи через HTTP"""
        wb = Workbook()
        wb.remove(wb.active)
        
        self.create_info_sheet(wb)
        self.create_raw_data_sheet(wb)
        self.create_expert_comparisons_sheet(wb)
        self.create_matrices_sheet(wb)
        self.create_fuzzy_sets_sheet(wb)
        self.create_weights_sheet(wb)
        self.create_step_by_step_sheet(wb)
        self.create_final_results_sheet(wb)
        self.create_consistency_sheet(wb)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output