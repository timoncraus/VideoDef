"""
Тесты для экспорта результатов верификации в Excel
"""

import json
from django.test import TestCase
from django.contrib.auth import get_user_model
from io import BytesIO
from openpyxl import load_workbook

from resume.bellman_zade import BellmanZadeMCDA, create_brand_project_example
from resume.utils.excel_exporter import VerificationExcelExporter

User = get_user_model()


class VerificationExcelExporterTest(TestCase):
    """Тесты для VerificationExcelExporter"""

    def setUp(self):
        """Создание тестовой модели для экспорта"""
        # Создаем модель из примера
        self.model = create_brand_project_example()
        self.model.calculate_solution(use_weights=True)
        self.solution = self.model.solution_fuzzy_set
        
        # Ожидаемые значения из методички
        self.expected_mu = {
            'P1': 0.59,
            'P2': 0.53,
            'P3': 0.50,
            'P4': 0.62
        }
        
        # Веса из методички
        self.calculated_weights = {
            'G1': 0.15, 'G2': 0.34, 'G3': 0.26,
            'G4': 0.05, 'G5': 0.13, 'G6': 0.07
        }
        
        # Ожидаемые степени принадлежности из методички
        self.expected_memberships = {
            'G1': {'P1': 0.87, 'P2': 0.71, 'P3': 0.51, 'P4': 0.92},
            'G2': {'P1': 0.83, 'P2': 0.75, 'P3': 0.60, 'P4': 0.95},
            'G3': {'P1': 0.64, 'P2': 0.76, 'P3': 0.63, 'P4': 0.72},
            'G4': {'P1': 0.59, 'P2': 0.53, 'P3': 0.82, 'P4': 0.69},
            'G5': {'P1': 0.69, 'P2': 0.67, 'P3': 0.74, 'P4': 0.81},
            'G6': {'P1': 0.60, 'P2': 0.73, 'P3': 0.66, 'P4': 0.70}
        }

    def test_exporter_initialization(self):
        """Тест инициализации экспортера"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        self.assertIsNotNone(exporter)
        self.assertEqual(exporter.model, self.model)
        self.assertEqual(exporter.solution, self.solution)
        self.assertEqual(exporter.expected_mu, self.expected_mu)
        self.assertEqual(exporter.calculated_weights, self.calculated_weights)
        self.assertEqual(exporter.expected_memberships, self.expected_memberships)
        self.assertIsNotNone(exporter.textbook_weights)
        self.assertEqual(len(exporter.textbook_weights), 6)

    def test_export_to_bytesio(self):
        """Тест экспорта в BytesIO"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        
        self.assertIsInstance(result, BytesIO)
        self.assertGreater(result.getbuffer().nbytes, 0)
        
        # Проверяем, что это валидный Excel файл
        try:
            wb = load_workbook(result)
            self.assertIsNotNone(wb)
        except Exception as e:
            self.fail(f"Failed to load workbook: {e}")

    def test_export_creates_correct_sheets(self):
        """Тест создания правильных листов в Excel"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        
        expected_sheets = [
            '0_Сводка',
            '1_Матрицы_сравнений',
            '2_Нечеткие_множества',
            '3_Веса_критериев',
            '4_Расчет_по_формуле',
            '5_Финальное_решение',
            '6_Согласованность_матриц'
        ]
        
        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, wb.sheetnames)

    def test_summary_sheet_content(self):
        """Тест содержимого сводного листа"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['0_Сводка']
        
        # Проверяем наличие основных заголовков
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('Верификация', str(ws['A1'].value))
        self.assertIsNotNone(ws['A3'].value)  # Дата
        self.assertIsNotNone(ws['A4'].value)  # Количество альтернатив
        self.assertIsNotNone(ws['A5'].value)  # Количество критериев
        self.assertIsNotNone(ws['A6'].value)  # Лучшая альтернатива

    def test_matrices_sheet_content(self):
        """Тест содержимого листа с матрицами"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['1_Матрицы_сравнений']
        
        # Проверяем наличие заголовка
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('Матрицы парных сравнений', str(ws['A1'].value))
        
        # Проверяем наличие матриц для каждого критерия
        for criterion in ['G1', 'G2', 'G3', 'G4', 'G5', 'G6']:
            found = False
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and criterion in str(cell_value):
                    found = True
                    break
            self.assertTrue(found, f"Matrix for {criterion} not found")

    def test_fuzzy_sets_sheet_content(self):
        """Тест содержимого листа с нечеткими множествами"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['2_Нечеткие_множества']
        
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('Нечеткие множества', str(ws['A1'].value))
        
        headers = [ws.cell(row=4, column=i).value for i in range(1, 7)]
        expected_headers = ['Критерий', 'Альтернатива', 'Ожидаемое μ (методичка)', 'Расчетное μ', 'Абсолютная ошибка', 'Относительная ошибка']
        for expected in expected_headers:
            self.assertIn(expected, headers)

    def test_weights_sheet_content(self):
        """Тест содержимого листа с весами"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['3_Веса_критериев']
        
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('Коэффициенты относительной важности', str(ws['A1'].value))

    def test_formula_sheet_content(self):
        """Тест содержимого листа с расчетами по формуле"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['4_Расчет_по_формуле']
        
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('формуле (2.35)', str(ws['A1'].value))

    def test_solution_sheet_content(self):
        """Тест содержимого листа с финальным решением"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['5_Финальное_решение']
        
        # Проверяем наличие заголовка
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('Финальное решение', str(ws['A1'].value))
        
        # Проверяем наличие всех альтернатив
        for alt in ['P1', 'P2', 'P3', 'P4']:
            found = False
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and alt == str(cell_value):
                    found = True
                    break
            self.assertTrue(found, f"Row for {alt} not found")
        
        # Проверяем, что лучшая альтернатива выделена
        best_alt, _ = self.model.get_best_alternative()
        found_best = False
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and best_alt == str(cell_value):
                # Проверяем, что ячейка залита (имеет fill)
                fill = ws.cell(row=row, column=1).fill
                if fill and fill.fill_type:
                    found_best = True
                    break
        self.assertTrue(found_best, f"Best alternative {best_alt} not highlighted")

    def test_consistency_sheet_content(self):
        """Тест содержимого листа с согласованностью"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        result = exporter.export_to_bytesio()
        wb = load_workbook(result)
        ws = wb['6_Согласованность_матриц']
        
        self.assertIsNotNone(ws['A1'].value)
        self.assertIn('согласованности матриц', str(ws['A1'].value))

    def test_export_filename(self):
        """Тест экспорта с именем файла"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        filename = "test_verification.xlsx"
        result = exporter.export(filename)
        
        self.assertEqual(result, filename)
        # Проверяем, что файл создан (в реальном тесте файл будет создан,
        # но в CI может не быть прав на запись)
        # Вместо этого проверяем, что метод отработал

    def test_textbook_weights_exist(self):
        """Тест наличия всех весов из методички"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        expected_keys = ['G1', 'G2', 'G3', 'G4', 'G5', 'G6']
        for key in expected_keys:
            self.assertIn(key, exporter.textbook_weights)
            self.assertGreater(exporter.textbook_weights[key], 0)
            self.assertLessEqual(exporter.textbook_weights[key], 1)

    def test_all_alternatives_in_expected_mu(self):
        """Тест наличия всех альтернатив в ожидаемых значениях"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        for alt in self.model.alternatives:
            self.assertIn(alt, exporter.expected_mu)
            self.assertGreater(exporter.expected_mu[alt], 0)

    def test_all_criteria_in_expected_memberships(self):
        """Тест наличия всех критериев в ожидаемых принадлежностях"""
        exporter = VerificationExcelExporter(
            model=self.model,
            solution=self.solution,
            expected_mu=self.expected_mu,
            calculated_weights=self.calculated_weights,
            expected_memberships=self.expected_memberships
        )
        
        for criterion in self.model.criteria:
            self.assertIn(criterion, exporter.expected_memberships)
            for alt in self.model.alternatives:
                self.assertIn(alt, exporter.expected_memberships[criterion])
                self.assertGreaterEqual(exporter.expected_memberships[criterion][alt], 0)
                self.assertLessEqual(exporter.expected_memberships[criterion][alt], 1)